from openpyxl import load_workbook
# Set the inital value 
wb = load_workbook("Sample_SQL진단결과보고서.xlsx")
tableName = "F100032818"
fileName = "대전광역시_중구_공동주택_20210827"

# Set the type of test : 'kor', 'num', 'truth'
col_list = {}
col_list[3] = 'num'
col_list[4] = 'num'
col_list[7] = 'date'

# [01.컬럼목록]에서 '테이블명', '파일명', 'sql 진단 여부'에 값 입력하기
ws1 = wb["01.컬럼목록"]
num_Col = ws1.max_row - 1
for i in range(2, ws1.max_row + 1):
    ws1.cell(column=2, row=i, value=tableName)
    ws1.cell(column=3, row=i, value=fileName)
    if i in col_list.keys():
        ws1.cell(column=6, row=i+1, value='Y')

# [Seet2]
ws2 = wb["Seet2"]
ws2.cell(column=2, row=1, value=tableName)
ws2.cell(column=2, row=2, value=fileName)

# test해야 하는 컬럼 찾기, 해당 컬럼에 대한 sheet 추가해주기
for x in range(2, num_Col + 2):
    if (ws1.cell(column=6, row= x).value == 'Y'): # x=2
        y  = str(x)
        y = wb.copy_worksheet(ws2)  # y = '2'

        col_Num = ws1.cell(column=4, row=x).value
        columName = ws1.cell(column=5, row=x).value
        testKind = col_list[x-1]
        content = ""
        sql = ""

        if testKind == 'num':
            content = columName + "은 숫자만 들어가야한다. 마이너스와 소수점 허용"
            sql = " '^[-]?\d+(\.?\d*)$'); "
        elif testKind == 'kor':
            content = columName + "의 값은 한글만 들어가야한다."
            sql = " '[가-힝]'); "
        elif testKind == 'date':
            content = columName + "의 값은 YYYY-MM-DD 데이터 유형을 따라야 한다."
            sql = " '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$'); "
        elif testKind == 'truth':
            content = columName + "의 값은 'Y', 'N' 중 하나여야 한다."
            sql = " and UPPER(" + col_Num + ") not in ('Y', 'N', '1', '0'); "
        else:
            content = columName + "의 값은 전화번호 형식이어야 한다."
            sql = " '^[0-9]{2,3}-[0-9]{3,4}-[0-9]{4}$'); "   

        y.cell(column=2, row=3, value=col_Num)
        y.cell(column=2, row=4, value=columName)
        y.cell(column=2, row=5, value=content)
        y.cell(column=2, row=6, value="select " + col_Num + " /* 컬럼명: " + columName + " */ \n" +
                                    "from C##OPENDATA." + tableName + " /* 파일명: " + fileName +  " */ \n" + 
                                    "where \"index\" <> 0 " + "/* index 0 번은 항목명*/" + "  \n" +
                                    "and not REGEXP_LIKE(" + col_Num + ", " + sql      
            )    
        y.title = col_Num
        

print(col_list)


# Delete the sheet name "C2"
wb.remove(ws2)

wb.save(fileName+ "_" +tableName +"_SQL진단결과보고서.xlsx")