from openpyxl import styles
from openpyxl.styles import Font, Border, Alignment
from openpyxl import load_workbook

from openpyxl.styles.borders import Side
import pandas as pd

# Set the inital value 
url = "C:\pythonWorkSpace"
wb = load_workbook("Sample_SQL진단결과보고서.xlsx")
tableName = "F100044656"
fileName = "충청남도_아산시_이용업현황_20190923"
sizeFileName = len(fileName)
print(sizeFileName)

# Set the type of test : 'kor', 'num', 'truth', 'day', 'time', daytime'
col_list = {}
col_list[4] = 'day'

col_list[9] = 'num'
col_list[10] = 'phone'
col_list[11] = 'day'
for i in range(14, 20):
    col_list[i] = 'num'


print(col_list)

# columns 목록 
df = pd.read_csv(tableName+'.csv')
columnNames = []
for i in df.columns:
    columnNames.append(i)
print("Columns = ", columnNames)
columnLength = len(columnNames)
print("Column length = ", columnLength)

# [01.컬럼목록] 스타일: '파일명' 컬럼 너비 조정
ws1 = wb["01.컬럼목록"]

ws1.column_dimensions["C"].width = sizeFileName * 2

# [01.컬럼목록]에서 '테이블명', '파일명', '항목명', 'sql 진단 여부'에 값 입력하기
for i in range(1, columnLength+1): # i=2, 3, ... ,
    index = str(i)
    print("index = ", index) 
    ws1.cell(column=1, row=i+1, value= index)
    ws1.cell(column=2, row=i+1, value= tableName)
    ws1.cell(column=3, row=i+1, value= fileName)
    ws1.cell(column=4, row=i+1, value= "C" + index)
    ws1.cell(column=5, row=i+1, value= columnNames[i-1])
    if i in col_list.keys():
        ws1.cell(column=6, row=i+1, value='Y')
        ws1.cell(column=7, row=i+1, value= df.iloc[:, i-1].count())
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), bottom=Side(style="thin"))
    font = Font(name="맑은 고딕")
    data = ws1[i+1]
    for cell in data:
        cell.font = font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", horizontal="center")
    # ws1.cell.border = thin_border

#
col_name = ws1["E"]
for cell in col_name:
    print(cell.value)

# [Seet2]
ws2 = wb["Seet2"]
ws2.cell(column=2, row=1, value=tableName)
ws2.cell(column=2, row=2, value=fileName)

# test해야 하는 컬럼 찾기, 해당 컬럼에 대한 sheet 추가해주기
for k in col_list.keys():
    y = str(k)  # y = '5'
    y = wb.copy_worksheet(ws2)
    row_info = ws1[k+1]
    print(row_info)
    
    testKind = col_list[k]
    col_Num = ws1.cell(column=4, row=k+1).value
    columName = ws1.cell(column=5, row=k+1).value
    content = ""
    sql = ""

    if testKind == 'num':
        content = columName + "은 숫자만 들어가야한다. 마이너스와 소수점 허용. "
        sql = " '^[-]?\d+(\.?\d*)$'); "
    elif testKind == 'kor':
        content = columName + "의 값은 한글만 들어가야한다."
        sql = " '[가-힝]'); "
    elif testKind == 'day':
        content = columName + "의 값은 YYYY-MM-DD 데이터 유형을 따라야 한다."
        sql = " '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])$'); "
    elif testKind == 'time':
        content = columName + "의 값은 HH-MM 데이터 유형을 따라야 한다"
        sql = " '^([1-9]|[01][0-9]|2[0-4])[:]([0-5][0-9])$'); "
    elif testKind == 'daytime':
        content = columName + "의 값은 YYYY-MM-DD HH-MM 데이터 유형을 따라야 한다." 
        sql = " '^(19|20)\d{2}-(0[1-9]|1[012])-(0[1-9]|[12][0-9]|3[0-1])\s([1-9]|[01][0-9]|2[0-4])[:]([0-5][0-9])$'); "
    elif testKind == 'truth':
        content = columName + "의 값은 'Y', 'N' 중 하나여야 한다."
        sql = " and UPPER(" + col_Num + ") not in ('Y', 'N', '1', '0'); "
    else:
        content = columName + "의 값은 전화번호 형식이어야 한다."
        sql = " '^[0-9]{2,3}-[0-9]{3,4}-[0-9]{4}$'); "

    y.cell(column=2, row=3, value=col_Num)
    y.cell(column=2, row=4, value=columName)
    y.cell(column=2, row=5, value=content)
    y.cell(column=2, row=6, value="select " + col_Num + " /* 컬럼명: " + columName + " */ \n" +
                                    "from C##OPENDATA." + tableName + " /* 파일명: " + fileName +  " */ \n" + 
                                    "where \"index\" <> 0 " + "/* index 0 번은 항목명*/" + "  \n" +
                                    "and not REGEXP_LIKE(" + col_Num + ", " + sql     
            )    
    y.title = col_Num




# Delete the sheet name "C2"
wb.remove(ws2)

wb.save(fileName+ "_" +tableName +"_SQL진단결과보고서.xlsx")