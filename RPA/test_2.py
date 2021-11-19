from openpyxl import load_workbook
wb = load_workbook("인천교통공사_1호선날짜별승하차인원(2018년)_F100027951_SQL진단결과보고서.xlsx")
tableName = "F100027952"
fileName = "인천교통공사_1호선날짜별승하차인원(2018년)_20190503"

ws1 = wb["01.컬럼목록"]

# for y in range(2, 29):
#     ws1.cell(column=2, row=y, value=tableName)
#     ws1.cell(column=3, row=y, value=fileName)
    
#  insert the value "test1" in B3
# ws1["B3"] = "test1"
# ws1.cell(column=2, row=3, value="test1")  

# ws2 = wb["C2"]
# ws2.cell(column=2, row=1, value=tableName)
# ws2.cell(column=2, row=2, value=fileName)

# # 컬럼 복사
# for y in range(3, 28):
#     x = str(y) 
#     x = wb.copy_worksheet(ws2)
#     testCol = "C" + str(y)
#     print(ws1.cell(column=5, row=1+y).value)
#     columName = ws1.cell(column=5, row=1+y).value
    
#     content = " 값은 숫자만 들어가야한다. 마이너스와 소수점 허용."
#     # insert
#     x.cell(column=2, row=3, value=testCol)
#     x.cell(column=2, row=4, value=columName)
#     x.cell(column=2, row=5, value=content)
#     x.title = "C" + str(y)

for y in range(3, 28):
    x = str(y)
    x = wb["C" + x]
    x.cell(column=2, row=1, value=tableName)
    x.cell(column=2, row=2, value=fileName)
    x.cell(column=2, row=6, value="select C" + str(y) + " /* " + ws1.cell(column=5, row=1+y).value+ " */ \n" +
                                    "from C##OPENDATA.F100027952 /* 인천교통공사_1호선날짜별승하차인원(2018년)_20190503	 */ \n" + 
                                    "where \"index\" <> 0 " +
                                    "and not REGEXP_LIKE(C" +str(y)+ ", '^[-]?\d+(\.?\d*)$');"     
            )    

wb.save("인천교통공사_1호선날짜별승하차인원(2018년)_F100027952_SQL진단결과보고서.xlsx")
